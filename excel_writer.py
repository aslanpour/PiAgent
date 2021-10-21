import threading
import datetime
from openpyxl import load_workbook
from openpyxl.utils import FORMULAE
from openpyxl.styles import Font, Color, PatternFill

def write_avg(rows_count, excel_file_path):
    
    try:
        #get workbook
        wb = load_workbook(filename = excel_file_path)
        #get worksheet
        sheet = wb['nodes']
    except Exception as e:
        return str(e)

    #last written row   index 
    max_row = sheet.max_row
    #current row index
    new_row = max_row + 1
    
    #name it
    sheet["A" + str(new_row)] = "avg"
    #for i in range (100):
    #    sheet["C" + str(new_row)].font = Font(bold=True)
    
    #get indexes
    first_row = str(max_row - (rows_count - 1))
    end_row = str(max_row)
    
    #use iter_cols() ?????
    for row in sheet.iter_cols(min_row=new_row, max_row=new_row, min_col=2, max_col=100):
        for cell in row:
            column_letter = cell.column_letter
            cell.value= "=AVERAGE(" + column_letter + first_row + ":" + column_letter + end_row +")"
            #bold it
            cell.font = Font(bold=True)
            #fill color
            cell.fill = PatternFill(fgColor="DCDCDC")
        
    #columnIds = ['C', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X']
    #calculate test duration avg 
    #for columnId in columnIds:
        #set value
        #sheet[columnId + str(new_row)] = "=AVERAGE(" + columnId + first_row + ":" + columnId + end_row +")"
        #bold it
        #sheet[columnId + str(new_row)].font = Font(bold=True)
        #fill color
        #sheet[columnId + str(new_row)].fill = PatternFill(fgColor="00FF00FF")
    
    #save
    wb.save(filename = excel_file_path)
    wb.close()

    return new_row

def write_scheduler(metrics, excel_file_path):
    lock = threading.Lock()
    with lock:
        try:
            wb = load_workbook(filename = excel_file_path)
            sheet = wb['scheduler']
        except Exception as e:
            return str(e)
        #freeze first column and row
        sheet.freeze_panes = "B2"
        
        #write new row to the end
        max_row = sheet.max_row
        print('max_row= ' + str(max_row))
        
        #add scheduler info
        down_counter = metrics["scheduler"]["down_counter"]
        print(down_counter)
        for key, value in down_counter.items():
            row = []
            #test name A
            row.extend([metrics["info"]["test_name"]])            
            #timestamp B
            current_time = str(datetime.datetime.now(datetime.timezone.utc).astimezone())
            row.extend([current_time])
        
            #node name C
            row.extend([key])
            #down_counter D
            row.extend([value])
            
            #sum placements E
            row.extend([metrics["scheduler"]["placements"]["sum"]])
            #placements per worker F..
            row.extend([metrics["scheduler"]["placements"]["workers"][key]])
            
            #placements per functions
            #add placmement per function
            per_functions = metrics["scheduler"]["placements"]["functions"]
            for key, value in per_functions.items():
                #function name
                row.extend([key])
                #function placements
                row.extend([value])
                
            #append this node info
            sheet.append(row)  
            
            #stylish
            max_row = sheet.max_row
            current_row = max_row
            
            #bold node name
            sheet["C" + str(current_row)].font = Font(bold=True)
            
                            
        #save
        wb.save(filename = excel_file_path)
        wb.close()
        
#write metrics
def write(metrics, excel_file_path):
    result = "null"
    current_row = 0
    
    #write scheduler
    if "scheduler" in metrics:
        try:
            write_scheduler(metrics, excel_file_path)
        except Exception as e:
            print('write_scheduler:' + str(e))
        
    lock = threading.Lock()
    with lock:
        try:
            wb = load_workbook(filename = excel_file_path)
            sheet = wb['nodes']
        except Exception as e:
            return str(e)
        #freeze first column and row
        sheet.freeze_panes = "B2"
        
            
        #write new row to the end
        max_row = sheet.max_row
        print('max_row= ' + str(max_row))
        row = []

        #test name A
        row.extend([metrics["info"]["test_name"]])
        
        
        #timestamp B
        current_time = str(datetime.datetime.now(datetime.timezone.utc).astimezone())
        row.extend([current_time])
        
        #info  C D E 
        row.extend([metrics["info"]["test_duration"],
                  metrics["info"]["test_started"],
                  metrics["info"]["test_finished"]])
        print('info')
        print(row)
        
        #node F G H I J K L M N O P Q
        row.extend([metrics["node"]["name"],
                  metrics["node"]["power_usage"],
                  metrics["node"]["down_time"]["minute"],
                  metrics["node"]["down_time"]["percent"],
                  metrics["node"]["cpu_usage"]["avg"],
                  metrics["node"]["cpu_usage"]["max"],
                  metrics["node"]["cpu_usage_up"]["avg"],
                  metrics["node"]["cpu_usage_up"]["max"],
                  metrics["node"]["memory_usage"]["avg"],
                  metrics["node"]["memory_usage"]["max"],
                  metrics["node"]["bw_usage"]["sent_mb"],
                  metrics["node"]["bw_usage"]["recv_mb"]])

        #apps R S T U V W X Y Z AA ... AX
        #if node has been LOAD_GENERATOR OR STANDALONE
        if "app_order" in metrics:
            #and has had enabled app
            if len(metrics["app_order"]):
                for app in metrics["app_order"]:

                    row.extend([app])
                    row.extend([metrics[app]["created"],
                              metrics[app]["sent"]["sum"],
                              metrics[app]["sent"]["percent"],
                              metrics[app]["code200"]["sum"],
                              metrics[app]["code200"]["percent"],
                              metrics[app]["code500"],
                              metrics[app]["code502"],
                              metrics[app]["code503"],
                              metrics[app]["code-1"],
                              metrics[app]["others"],
                              metrics[app]["dropped"]["sum"],
                              metrics[app]["dropped"]["percent"],
                              metrics[app]["dropped_in_bootup"]["sum"],
                              metrics[app]["dropped_in_bootup"]["percent"],
                              metrics[app]["admission_dur"]["avg"],
                              metrics[app]["admission_dur"]["max"],
                              metrics[app]["queue_dur"]["avg"],
                              metrics[app]["queue_dur"]["max"],
                              metrics[app]["exec_dur"]["avg"],
                              metrics[app]["exec_dur"]["max"],
                              metrics[app]["round_trip"]["avg"],
                              metrics[app]["round_trip"]["max"],
                              metrics[app]["useless_exec_dur"],
                              metrics[app]["throughput2"]["avg"],
                              metrics[app]["percentiles"]["p0"],
                              metrics[app]["percentiles"]["p25"],
                              metrics[app]["percentiles"]["p50"],
                              metrics[app]["percentiles"]["p75"],
                              metrics[app]["percentiles"]["p90"],
                              metrics[app]["percentiles"]["p95"],
                              metrics[app]["percentiles"]["p99"],
                              metrics[app]["percentiles"]["p99.9"],
                              metrics[app]["percentiles"]["p100"]])

        
        #append to excel
        sheet.append(row)
        
        
        #stylish the new row
        current_row = max_row + 1
        #bold test name
        sheet["A" + str(current_row)].font = Font(bold=True)
        #bold node_name F
        sheet["F" + str(current_row)].font = Font(bold=True)
        #bold if not master
        if not "master" in metrics["node"]["name"]:
            #down_time percent 
            sheet["I" + str(current_row)].font = Font(bold=True)
            #cpuUtil_up_avg
            sheet["L" + str(current_row)].font = Font(bold=True)
            
        #save
        wb.save(filename = excel_file_path)
        wb.close()
        
        return 'success', current_row
